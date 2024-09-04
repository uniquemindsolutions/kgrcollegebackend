from django.shortcuts import render, get_object_or_404, redirect
from .models import *
from .forms import * 
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login as auth_login, logout as auth_logout, authenticate
from django.views.decorators.csrf import csrf_exempt
from .serializers import *
from django.http import JsonResponse
from rest_framework import viewsets
from .filters import * 
from django.core.mail import send_mail
from django.conf import settings
import json
from django.core.mail import EmailMessage
from django_filters.rest_framework import DjangoFilterBackend
import pandas as pd
import openpyxl
from django.http import HttpResponse
from django.http import FileResponse



def home(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            if user.is_staff or user.is_superuser:
                auth_login(request, user)
                return redirect('banner_list')
            else:
                messages.error(request, 'You do not have permission to access this area.')
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = AuthenticationForm()

    return render(request, 'home.html', {'form': form})

def logout(request):
    auth_logout(request)
    return redirect('login')

@login_required
@csrf_exempt
def dashboard(request):
    return render(request, 'dashboard/dashboard.html')


@login_required
@csrf_exempt
def banner_list(request):
    banners = Banner.objects.all().order_by('id')
    if request.headers.get('Content-Type') == 'application/json' or request.headers.get('Accept') == 'application/json':
        banner_data = serializers.serialize('json', banners)
        return JsonResponse(banner_data, safe=False)
    return render(request, 'Home_Page/banner_list.html', {'banners': banners})


@login_required
def banner_add(request):
    if request.method == 'POST':
        form = BannerForm(request.POST, request.FILES)
        if form.is_valid():
            banner = form.save(commit=False)
            image_file = banner.image

            # Open the image file using PIL to extract attributes
            with Image.open(image_file) as img:
                banner.width = img.width
                banner.height = img.height
                banner.format = img.format if img.format else 'Unknown'

            banner.file_size = image_file.size
            banner.save()
            return redirect('banner_list')
    else:
        form = BannerForm()
    return render(request, 'Home_Page/banner_form.html', {'form': form})

@login_required
def banner_edit(request, pk):
    banner = get_object_or_404(Banner, pk=pk)
    if request.method == "POST":
        form = BannerForm(request.POST, request.FILES, instance=banner)
        if form.is_valid():
            form.save()
            return redirect('banner_list')
    else:
        form = BannerForm(instance=banner)
    return render(request, 'Home_Page/banner_form.html', {'form': form})

@login_required
def banner_delete(request, pk):
    banner = get_object_or_404(Banner, pk=pk)
    banner.delete()
    return redirect('banner_list')

@csrf_exempt
def college_update_list_view(request):
    updates = CollegeUpdates.objects.all().order_by('id')
    return render(request, 'Home_Page/college_update_list.html', {'updates': updates})

@csrf_exempt
def college_update_create_view(request):
    if request.method == "POST":
        form = CollegeUpdatesForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('college_update_list')
    else:
        form = CollegeUpdatesForm()
    return render(request, 'Home_Page/college_update_form.html', {'form': form})

@csrf_exempt
def college_update_edit_view(request, id):
    update = get_object_or_404(CollegeUpdates, id=id)
    if request.method == "POST":
        form = CollegeUpdatesForm(request.POST, instance=update)
        if form.is_valid():
            form.save()
            return redirect('college_update_list')
    else:
        form = CollegeUpdatesForm(instance=update)
    return render(request, 'Home_Page/college_update_form.html', {'form': form})

@csrf_exempt
def college_update_delete_view(request, id):
    update = get_object_or_404(CollegeUpdates, id=id)
    if request.method == "POST":
        update.delete()
        return redirect('college_update_list')
    return render(request, 'Home_Page/college_update_list.html')

    
@csrf_exempt
def student_count_list_view(request):
    counts = StudentsCount.objects.all()
    return render(request, 'Home_Page/student_count_list.html', {'counts': counts})

@csrf_exempt
def student_count_create_view(request):
    if request.method == "POST":
        form = StudentsCountForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('student_count_list')
    else:
        form = StudentsCountForm()
    return render(request, 'Home_Page/student_count_form.html', {'form': form})

@csrf_exempt
def student_count_edit_view(request, id):
    count = get_object_or_404(StudentsCount, id=id)
    if request.method == "POST":
        form = StudentsCountForm(request.POST, instance=count)
        if form.is_valid():
            form.save()
            return redirect('student_count_list')
    else:
        form = StudentsCountForm(instance=count)
    return render(request, 'Home_Page/student_count_form.html', {'form': form})

@csrf_exempt
def student_count_delete_view(request, id):
    count = get_object_or_404(StudentsCount, id=id)
    if request.method == "POST":
        count.delete()
        return redirect('student_count_list')
    return redirect('student_count_list')

@csrf_exempt
def faculty_count_list_view(request):
    counts = FactulyCount.objects.all().order_by('id')
    return render(request, 'Home_Page/faculty_count_list.html', {'counts': counts})

@csrf_exempt
def faculty_count_create_view(request):
    if request.method == "POST":
        form = FacultyCountForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('faculty_count_list')
    else:
        form = FacultyCountForm()
    return render(request, 'Home_Page/faculty_count_form.html', {'form': form})

@csrf_exempt
def faculty_count_edit_view(request, id):
    count = get_object_or_404(FactulyCount, id=id)
    if request.method == "POST":
        form = FacultyCountForm(request.POST, instance=count)
        if form.is_valid():
            form.save()
            return redirect('faculty_count_list')
    else:
        form = FacultyCountForm(instance=count)
    return render(request, 'Home_Page/faculty_count_form.html', {'form': form})

@csrf_exempt
def faculty_count_delete_view(request, id):
    count = get_object_or_404(FactulyCount, id=id)
    if request.method == "POST":
        count.delete()
        return redirect('faculty_count_list')
    return redirect('faculty_count_list')

@csrf_exempt
def programs_count_list_view(request):
    counts = ProgramsCount.objects.all().order_by('id')
    return render(request, 'Home_Page/programs_count_list.html', {'counts': counts})

@csrf_exempt
def programs_count_create_view(request):
    if request.method == "POST":
        form = ProgramsCountForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('programs_count_list')
    else:
        form = ProgramsCountForm()
    return render(request, 'Home_Page/programs_count_form.html', {'form': form})

@csrf_exempt
def programs_count_edit_view(request, id):
    count = get_object_or_404(ProgramsCount, id=id)
    if request.method == "POST":
        form = ProgramsCountForm(request.POST, instance=count)
        if form.is_valid():
            form.save()
            return redirect('programs_count_list')
    else:
        form = ProgramsCountForm(instance=count)
    return render(request, 'Home_Page/programs_count_form.html', {'form': form})

@csrf_exempt
def programs_count_delete_view(request, id):
    count = get_object_or_404(ProgramsCount, id=id)
    if request.method == "POST":
        count.delete()
        return redirect('programs_count_list')
    return redirect('programs_count_list')

def student_form_view(request):
    if request.method == "POST":
        form = StudentFormForm(request.POST)
        if form.is_valid():
            student_form = form.save()
            # Send email
            send_mail(
                'New Student Form Submission',
                f"Name: {student_form.name}\nEmail: {student_form.email}\nPhone Number: {student_form.phone_number}",
                settings.DEFAULT_FROM_EMAIL,
                ['cvamshikrishna9381@gmail.com'],
                fail_silently=False,
            )
            return redirect('success_page')  # Redirect to a success page or another view
    else:
        form = StudentFormForm()
    return render(request, 'Home_Page/student_form.html', {'form': form})

@csrf_exempt
def send_form_email(request):
    if request.method == "POST":
        try:
            # Parse JSON data from the request body
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)

        # Extract the fields from the data
        name = data.get('name')
        email = data.get('email')
        phone = data.get('phone')

        # Check for missing fields
        if not name or not email or not phone:
            return JsonResponse({'error': 'Missing required fields'}, status=400)

        # Send email
        try:
            send_mail(
                'New Contact Form Submission',
                f"Name: {name}\nEmail: {email}\nPhone: {phone}",
                settings.DEFAULT_FROM_EMAIL,
                ['cvamshikrishna9381@gmail.com'],  # Your recipient email
                fail_silently=False,
            )
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

        # Respond with success
        return JsonResponse({'message': 'Email sent successfully'}, status=200)

    return JsonResponse({'error': 'Invalid request'}, status=400)

def student_form_list_view(request):
    forms = StudentForm.objects.all().order_by('id')
    return render(request, 'Home_Page/student_form_list.html', {'forms': forms})

def success_page(request):
    return render(request, 'Home_Page/success_page.html')


@csrf_exempt
def faculty_mba_list_view(request):
    faculties = Faculty_Mba.objects.all().order_by('id')
    return render(request, 'Faculty_Page/faculty_mba_list.html', {'faculties': faculties})

@csrf_exempt
def faculty_mba_create_view(request):
    if request.method == 'POST':
        form = FacultyMBAForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('faculty_mba_list')
    else:
        form = FacultyMBAForm()
    return render(request, 'Faculty_Page/faculty_mba_form.html', {'form': form})

@csrf_exempt
def bulk_upload_mba_faculty(request):
    bulk_form = BulkFacultyUploadForm(request.POST, request.FILES or None)

    if request.method == 'POST' and bulk_form.is_valid():
        excel_file = request.FILES['excel_file']

        try:
            # Load the Excel file
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Iterate through the rows in the sheet and create Faculty_MBA objects
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, designation, qualification, experience_teaching = row

                # Ensure slno is not null
                Faculty_Mba.objects.create(
                    name=name,
                    designation=designation,
                    qualification=qualification,
                    experience_teaching=experience_teaching
                )

            messages.success(request, "MBA Faculty members uploaded successfully!")
            return redirect('faculty_mba_list')
        except Exception as e:
            messages.error(request, f"Error during file upload: {e}")

    return render(request, 'Faculty_Page/bulk_upload_mba_faculty.html', {
        'bulk_form': bulk_form
    })

@csrf_exempt
def faculty_mba_edit_view(request, id):
    faculty = get_object_or_404(Faculty_Mba, id=id)
    if request.method == 'POST':
        form = FacultyMBAForm(request.POST, instance=faculty)
        if form.is_valid():
            form.save()
            return redirect('faculty_mba_list')
    else:
        form = FacultyMBAForm(instance=faculty)
    return render(request, 'Faculty_Page/faculty_mba_form.html', {'form': form})

@csrf_exempt
def faculty_mba_delete_view(request, id):
    faculty = get_object_or_404(Faculty_Mba, id=id)
    if request.method == 'POST':
        faculty.delete()
        return redirect('faculty_mba_list')
    return render(request, 'Faculty_Page/faculty_mba_list.html')


@csrf_exempt
def faculty_pharmacy_list_view(request):
    faculties = Faculty_Pharamacy.objects.all().order_by('id')
    return render(request, 'Faculty_Page/faculty_pharmacy_list.html', {'faculties': faculties})

@csrf_exempt
def faculty_pharmacy_create_view(request):
    if request.method == 'POST':
        form = FacultyPharmacyForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('faculty_pharmacy_list')
    else:
        form = FacultyPharmacyForm()
    return render(request, 'Faculty_Page/faculty_pharmacy_form.html', {'form': form})

@csrf_exempt
def bulk_upload_pharamacy_faculty(request):
    bulk_form = BulkFacultyUploadForm(request.POST, request.FILES or None)

    if request.method == 'POST' and bulk_form.is_valid():
        excel_file = request.FILES['excel_file']

        try:
            # Load the Excel file
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            # Iterate through the rows in the sheet and create Faculty_MBA objects
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, designation, qualification, experience_teaching = row

                # Ensure slno is not null
                Faculty_Pharamacy.objects.create(
                    name=name,
                    designation=designation,
                    qualification=qualification,
                    experience_teaching=experience_teaching
                )

            messages.success(request, "Pharamacy Faculty members uploaded successfully!")
            return redirect('faculty_mba_list')
        except Exception as e:
            messages.error(request, f"Error during file upload: {e}")

    return render(request, 'Faculty_Page/bulk_upload_pharamacy_faculty.html', {
        'bulk_form': bulk_form
    })

@csrf_exempt
def faculty_pharmacy_edit_view(request, id):
    faculty = get_object_or_404(Faculty_Pharamacy, id=id)
    if request.method == 'POST':
        form = FacultyPharmacyForm(request.POST, instance=faculty)
        if form.is_valid():
            form.save()
            return redirect('faculty_pharmacy_list')
    else:
        form = FacultyPharmacyForm(instance=faculty)
    return render(request, 'Faculty_Page/faculty_pharmacy_form.html', {'form': form})

@csrf_exempt
def faculty_pharmacy_delete_view(request, id):
    faculty = get_object_or_404(Faculty_Pharamacy, id=id)
    if request.method == 'POST':
        faculty.delete()
        return redirect('faculty_pharmacy_list')
    return render(request, 'Faculty_Page/faculty_pharmacy_list.html')

# Gallery Images Views
def gallery_image_list(request):
    gallery_images = GalleryImages.objects.all().order_by('id')
    return render(request, 'Gallery_Page/gallery_images_list.html', {'gallery_images': gallery_images})

def gallery_image_create(request):
    if request.method == 'POST':
        form = GalleryImagesForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('gallery_images_list')
    else:
        form = GalleryImagesForm()
    return render(request, 'Gallery_Page/gallery_images_form.html', {'form': form})

def gallery_image_edit(request, pk):
    image = get_object_or_404(GalleryImages, pk=pk)
    if request.method == 'POST':
        form = GalleryImagesForm(request.POST, request.FILES, instance=image)
        if form.is_valid():
            form.save()
            return redirect('gallery_images_list')
    else:
        form = GalleryImagesForm(instance=image)
    return render(request, 'Gallery_Page/gallery_images_form.html', {'form': form})

def gallery_image_delete(request, pk):
    image = get_object_or_404(GalleryImages, pk=pk)
    if request.method == 'POST':
        image.delete()
        return redirect('gallery_images_list')
    return redirect('gallery_images_list')

# Gallery Videos Views
def gallery_video_list(request):
    gallery_videos = GalleryVideos.objects.all().order_by('id')
    return render(request, 'Gallery_Page/gallery_videos_list.html', {'gallery_videos': gallery_videos})

def gallery_video_create(request):
    if request.method == 'POST':
        form = GalleryVideosForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('gallery_videos_list')
    else:
        form = GalleryVideosForm()
    return render(request, 'Gallery_Page/gallery_videos_form.html', {'form': form})

def gallery_video_edit(request, pk):
    video = get_object_or_404(GalleryVideos, pk=pk)
    if request.method == 'POST':
        form = GalleryVideosForm(request.POST, request.FILES, instance=video)
        if form.is_valid():
            form.save()
            return redirect('gallery_videos_list')
    else:
        form = GalleryVideosForm(instance=video)
    return render(request, 'Gallery_Page/gallery_videos_form.html', {'form': form})

def gallery_video_delete(request, pk):
    video = get_object_or_404(GalleryVideos, pk=pk)
    if request.method == 'POST':
        video.delete()
        return redirect('gallery_videos_list')
    return redirect('gallery_videos_list')

def alumni_list(request):
    # Retrieve all alumni entries ordered by ID
    alumni_list = Alumni.objects.all().order_by('id')
    return render(request, 'Alumni_Page/Alumni_list.html', {'alumni_list': alumni_list})

def alumni_create(request):
    # Create a new alumni entry
    if request.method == 'POST':
        form = AlumniForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('alumni_list')
    else:
        form = AlumniForm()
    return render(request, 'Alumni_Page/Alumni_Form.html', {'form': form})

def alumni_edit(request, pk):
    # Edit an existing alumni entry
    alumni = get_object_or_404(Alumni, pk=pk)
    if request.method == 'POST':
        form = AlumniForm(request.POST, request.FILES, instance=alumni)
        if form.is_valid():
            form.save()
            return redirect('alumni_list')
    else:
        form = AlumniForm(instance=alumni)
    return render(request, 'Alumni_Page/Alumni_Form.html', {'form': form})

def alumni_delete(request, pk):
    # Delete an existing alumni entry
    alumni = get_object_or_404(Alumni, pk=pk)
    if request.method == 'POST':
        alumni.delete()
        return redirect('alumni_list')
    return redirect('alumni_list')

def events_list(request):
    events_list = EventsandActivites.objects.all().order_by('id')
    return render(request, 'Event_and_Activites_page/Events_and_Activites_list.html', {'events_list': events_list})

def events_create(request):
    if request.method == 'POST':
        form = EventsandActivitesForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('events_list')
    else:
        form = EventsandActivitesForm()
    return render(request, 'Event_and_Activites_page/Events_and_Activites_form.html', {'form': form})

def events_edit(request, id):
    event = get_object_or_404(EventsandActivites, id=id)
    if request.method == 'POST':
        form = EventsandActivitesForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            form.save()
            return redirect('events_list')
    else:
        form = EventsandActivitesForm(instance=event)
    return render(request, 'Event_and_Activites_page/Events_and_Activites_form.html', {'form': form})

def events_delete(request, id):
    event = get_object_or_404(EventsandActivites, id=id)
    if request.method == 'POST':
        event.delete()
        return redirect('events_list')
    return redirect('events_list')


@csrf_exempt
def send_registration_email(request):
    if request.method == 'POST':
        try:
            # Extract the form data
            student_name = request.POST.get('studentName')
            father_name = request.POST.get('fatherName')
            dob = request.POST.get('dob')
            address = request.POST.get('address')
            contact = request.POST.get('contact')
            email = request.POST.get('email')
            mba_status = request.POST.get('mba')
            passed_year = request.POST.get('passedYear')
            college = request.POST.get('college')
            university = request.POST.get('university')
            percentage = request.POST.get('percentage')

            # Handle file upload
            uploaded_file = request.FILES.get('upload')
            
            # Construct the email message
            message = f"""
            Student Name: {student_name}
            Father's Name: {father_name}
            Date of Birth: {dob}
            Permanent Address: {address}
            Contact Number: {contact}
            Email Address: {email}
            MBA Status: {mba_status}
            Passed Year: {passed_year if passed_year else 'N/A'}
            College: {college}
            University: {university}
            Percentage of Marks: {percentage}
            File Uploaded: {uploaded_file.name if uploaded_file else 'No file uploaded'}
            """

            # Create an EmailMessage object
            email_message = EmailMessage(
                subject="Student Registration Details",
                body=message,
                from_email=email,
                to=["cvamshikrishna9381@gmail.com"],
            )

            # Attach the uploaded file, if any
            if uploaded_file:
                email_message.attach(uploaded_file.name, uploaded_file.read(), uploaded_file.content_type)

            # Send the email
            email_message.send(fail_silently=False)

            return JsonResponse({"status": "success", "message": "Email sent successfully!"}, status=200)

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

    return JsonResponse({"status": "error", "message": "Invalid request method"}, status=405)

# View to handle file upload
def committees_view(request):
    if request.method == 'POST':
        form = CommitteesForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('committees_list')
    else:
        form = CommitteesForm()
    return render(request, 'Committees/committees_file_upload.html', {'form': form})

# View to list uploaded files
def committees_list_view(request):
    files_list = Committees.objects.all()  # Fetch all files
    return render(request, 'Committees/committees_list.html', {'files_list': files_list})  # Pass the files to the template


# View to handle file update
def committees_update_view(request, pk):
    file_obj = get_object_or_404(Committees, pk=pk)
    if request.method == 'POST':
        form = CommitteesForm(request.POST, request.FILES, instance=file_obj)
        if form.is_valid():
            form.save()
            return redirect('committees_list')
    else:
        form = CommitteesForm(instance=file_obj)
    return render(request, 'Committees/committees_file_upload.html', {'form': form})

# View to handle file delete
def committees_delete_view(request, pk):
    file_obj = get_object_or_404(Committees, pk=pk)
    file_obj.delete()
    return redirect('committees_list')
 
class BannerViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Banner.objects.all().order_by('id')
    serializer_class = BannerSerializer
    filter_backends = [Bannerfilter]
    
class CollegeUpdatesViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CollegeUpdates.objects.all().order_by('id')
    serializer_class = CollegeUpdatesSerializer

class StudentsCountViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = StudentsCount.objects.all().order_by('id')
    serializer_class = StudentsCountSerializer
    

class FactulyCountViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = FactulyCount.objects.all().order_by('id')
    serializer_class = FactulyCountSerializer

class ProgramsCountViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ProgramsCount.objects.all().order_by('id')
    serializer_class = ProgramsCountSerializer
   

class StudentFormViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = StudentForm.objects.all().order_by('id')
    serializer_class = StudentFormSerializer
    

class GalleryImagesViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = GalleryImages.objects.all().order_by('id')
    serializer_class = GalleryImagesSerializer
    filter_backends = [GalleryImagesfilter]

class GalleryVideosViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = GalleryVideos.objects.all().order_by('id')
    serializer_class = GalleryVideosSerializer
    filter_backends = [GalleryVideosfilter]

class Faculty_MbaViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Faculty_Mba.objects.all().order_by('id')
    serializer_class = Faculty_MbaSerializer
    

class Faculty_PharamacyViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Faculty_Pharamacy.objects.all().order_by('id')
    serializer_class = Faculty_PharamacySerializer
   

class AlumniViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Alumni.objects.all().order_by('id')
    serializer_class = AlumniSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = AlumniFilter  # Use custom filter set

class EventsandActivitesViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = EventsandActivites.objects.all().order_by('id')
    serializer_class = EventsandActivitesSerializer
    filter_backends = [EventsandActivitesfilter]

class CommitteesViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Committees.objects.all().order_by('id')
    serializer_class = CommitteesSerializer